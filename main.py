from fpdf import FPDF
import pandas as pd
import glob
import openpyxl
from pathlib import Path


filepaths = glob.glob('invoices/*.xlsx')

for file in filepaths:
    df = pd.read_excel(file, sheet_name='Sheet 1')
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=False, margin=0)
    pdf.add_page()
# Header
    filename = Path(file).stem
    invoice_nr = filename.split('-')[0]
    invoice_date = filename.split('-')[1]

    pdf.set_font(family='Times', style='B', size=16)
    pdf.cell(w=50, h=12, txt=f"Invoice nr. {invoice_nr}", align='L', ln=1)
    pdf.cell(w=50, h=12, txt=f"Date {invoice_date}", align='L', ln=1)
# Table Header
    header_columns = df.columns
    header_columns = [item.replace('_', ' ').title() for item in header_columns]
    pdf.set_font(family='Times', size=10, style='B')
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt=f"{header_columns[0]}",
             align='L', border=1)
    pdf.cell(w=70, h=8, txt=f"{header_columns[1]}",
             align='L', border=1)
    pdf.cell(w=35, h=8, txt=f"{header_columns[2]}",
             align='L', border=1)
    pdf.cell(w=30, h=8, txt=f"{header_columns[3]}",
             align='L', border=1)
    pdf.cell(w=30, h=8, txt=f"{header_columns[4]}",
             align='L', border=1, ln=1)
# Table Rows
    for i, row in df.iterrows():
        pdf.set_font(family='Times', size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, txt=f"{row['product_id']}",
                 align='L', border=1)
        pdf.cell(w=70, h=8, txt=f"{row['product_name']}",
                 align='L', border=1)
        pdf.cell(w=35, h=8, txt=f"{row['amount_purchased']}",
                 align='L', border=1)
        pdf.cell(w=30, h=8, txt=f"{row['price_per_unit']}",
                 align='L', border=1)
        pdf.cell(w=30, h=8, txt=f"{row['total_price']}",
                 align='L', border=1, ln=1)
# Total Price Row In Table
    totalprice = df['total_price'].sum()
    print(totalprice)
    pdf.set_font(family='Times', size=10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt=f"", align='L', border=1)
    pdf.cell(w=70, h=8, txt=f"", align='L', border=1)
    pdf.cell(w=35, h=8, txt=f"", align='L', border=1)
    pdf.cell(w=30, h=8, txt=f"", align='L', border=1)
    pdf.cell(w=30, h=8, txt=f"{totalprice}", align='L', border=1, ln=1)
# Total Price In Doc
    pdf.set_font(family='Times', size=10, style='B')
    pdf.cell(w=30, h=8, txt=f"Total Price is {totalprice}",
             align='L', ln=1)
# Company Name and Logo
    pdf.set_font(family='Times', size=10, style='B')
    pdf.cell(w=30, h=12, txt=f"@CompanyName",
             align='L')
    pdf.image('images/python.png', w=12)



    pdf.output(f'PDF/{filename}.pdf')







